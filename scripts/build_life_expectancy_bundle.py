#!/usr/bin/env python3
"""
Build DeathClock/Resources/life-expectancy-data.json from:
  - CDC US state period life tables (default: NVSR vol 74 no 12, 2022) or legacy LEWK4 (1999–2001)
  - World Bank male/female life expectancy at birth by country

Run from repo root:  python3 scripts/build_life_expectancy_bundle.py
"""

from __future__ import annotations

import argparse
import json
import re
import ssl
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
  import certifi
except ImportError:
  certifi = None

try:
  from openpyxl import load_workbook
except ImportError:
  print('Missing openpyxl. Run: pip install -r scripts/requirements.txt', file=sys.stderr)
  sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = REPO_ROOT / 'DeathClock' / 'Resources' / 'life-expectancy-data.json'

# LEWK4 (decennial 1999–2001)
CDC_LEWK4_INDEX = 'https://ftp.cdc.gov/pub/Health_Statistics/NCHS/Publications/NVSR/60_09/'
# NVSR 74-12: U.S. State Life Tables, 2022 — {ST}1=total, {ST}2=male, {ST}3=female, {ST}4=std errors
CDC_NVSR_74_12_INDEX = 'https://ftp.cdc.gov/pub/health_statistics/nchs/Publications/NVSR/74-12/'

WB_YEAR = 2022
WB_MALE = f'https://api.worldbank.org/v2/country/all/indicator/SP.DYN.LE00.MA.IN?format=json&date={WB_YEAR}:{WB_YEAR}&per_page=20000'
WB_FEMALE = f'https://api.worldbank.org/v2/country/all/indicator/SP.DYN.LE00.FE.IN?format=json&date={WB_YEAR}:{WB_YEAR}&per_page=20000'

US_SOURCE_NVSR = 'nvsr74-12'
US_SOURCE_LEWK4 = 'lewk4'

# USPS code -> display name (for settings picker)
US_STATE_NAMES: dict[str, str] = {
  'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas', 'CA': 'California',
  'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware', 'DC': 'District of Columbia',
  'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho', 'IL': 'Illinois',
  'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas', 'KY': 'Kentucky', 'LA': 'Louisiana',
  'ME': 'Maine', 'MD': 'Maryland', 'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota',
  'MS': 'Mississippi', 'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
  'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
  'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma', 'OR': 'Oregon',
  'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina', 'SD': 'South Dakota',
  'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah', 'VT': 'Vermont', 'VA': 'Virginia',
  'WA': 'Washington', 'WV': 'West Virginia', 'WI': 'Wisconsin', 'WY': 'Wyoming',
}

SLUG_TO_CODE: dict[str, str] = {name.lower().replace(' ', '_'): code for code, name in US_STATE_NAMES.items()}
SLUG_TO_CODE['district_of_columbia'] = 'DC'


def ssl_context() -> ssl.SSLContext:
  if certifi:
    return ssl.create_default_context(cafile=certifi.where())
  return ssl.create_default_context()


def http_get(url: str) -> bytes:
  req = urllib.request.Request(url, headers={'User-Agent': 'death-clock-bundle/1.0'})
  with urllib.request.urlopen(req, context=ssl_context(), timeout=120) as resp:
    return resp.read()


def parse_age_start(cell) -> int | None:
  if cell is None:
    return None
  s = str(cell).strip()
  # NVSR uses Unicode en dash (U+2013) in "0–1"; LEWK4 may use ASCII hyphen.
  s = s.replace('\u2013', '-').replace('\u2014', '-').replace('\u2212', '-')
  if re.match(r'^\d+\s*-\s*\d+', s):
    return int(s.split('-')[0].strip())
  if s.isdigit():
    return int(s)
  return None


def sheet_ex_series(ws, ex_col: int = 6) -> dict[int, float]:
  """Map exact age x -> expectation of life e_x from a life table sheet."""
  out: dict[int, float] = {}
  for row in ws.iter_rows(min_row=1, values_only=True):
    if not row:
      continue
    age = parse_age_start(row[0])
    if age is None:
      continue
    if ex_col >= len(row):
      continue
    ex = row[ex_col]
    if ex is None:
      continue
    try:
      out[age] = float(ex)
    except (TypeError, ValueError):
      continue
  return out


def dense_curve(points: dict[int, float], max_age: int = 100) -> list[float]:
  if not points:
    raise ValueError('empty life table')
  arr: list[float] = []
  last = points[0]
  for a in range(max_age + 1):
    if a in points:
      last = points[a]
    arr.append(last)
  return arr


def average_curves(curves: list[list[float]]) -> list[float]:
  if not curves:
    raise ValueError('no curves')
  n = min(len(c) for c in curves)
  return [sum(c[i] for c in curves) / len(curves) for i in range(n)]


# --- NVSR 74-12 (2022 state tables): one workbook per sex/total ---

def parse_nvsr_workbook(path: Path) -> list[float]:
  wb = load_workbook(path, data_only=True, read_only=True)
  try:
    ws = wb[wb.sheetnames[0]]
    return dense_curve(sheet_ex_series(ws))
  finally:
    wb.close()


def collect_nvsr_state_tables(cdc_dir: Path) -> dict[str, dict[str, list[float]]]:
  """Group CA1.xlsx / CA2.xlsx / CA3.xlsx -> male/female/total curves."""
  groups: dict[str, dict[int, Path]] = defaultdict(dict)
  for path in cdc_dir.iterdir():
    if not path.is_file():
      continue
    m = re.match(r'^([A-Za-z]{2})([1-4])\.xlsx$', path.name, flags=re.I)
    if not m:
      continue
    code = m.group(1).upper()
    num = int(m.group(2))
    if num == 4:
      continue
    groups[code][num] = path

  by_code: dict[str, dict[str, list[float]]] = {}
  for code in sorted(groups.keys()):
    if code not in US_STATE_NAMES:
      print(f'  skip unknown jurisdiction {code}', file=sys.stderr)
      continue
    parts = groups[code]
    if not all(i in parts for i in (1, 2, 3)):
      print(f'  skip {code}: missing 1/2/3 workbooks (have {sorted(parts.keys())})', file=sys.stderr)
      continue
    by_code[code] = {
      'total': parse_nvsr_workbook(parts[1]),
      'male': parse_nvsr_workbook(parts[2]),
      'female': parse_nvsr_workbook(parts[3]),
    }
  return by_code


def list_cdc_nvsr_xlsx() -> list[str]:
  html = http_get(CDC_NVSR_74_12_INDEX).decode('utf-8', errors='replace')
  found = re.findall(r'/74-12/([A-Za-z]{2}[1-4]\.xlsx)', html, flags=re.I)
  out: list[str] = []
  seen: set[str] = set()
  for f in found:
    m = re.match(r'^([A-Za-z]{2})([1-4])(\.xlsx)$', f, flags=re.I)
    if not m:
      continue
    norm = f'{m.group(1).upper()}{m.group(2)}{m.group(3).lower()}'
    key = norm.upper()
    if key in seen:
      continue
    seen.add(key)
    out.append(norm)
  return sorted(out)


def download_cdc_nvsr(cache_dir: Path) -> None:
  cache_dir.mkdir(parents=True, exist_ok=True)
  names = [n for n in list_cdc_nvsr_xlsx() if not re.search(r'4\.xlsx$', n, flags=re.I)]
  print(f'CDC NVSR 74-12: {len(names)} workbooks to fetch (skipping *4 standard errors)')
  for name in names:
    dest = cache_dir / name
    if dest.exists() and dest.stat().st_size > 1000:
      print(f'  cached {name}')
      continue
    url = CDC_NVSR_74_12_INDEX.rstrip('/') + '/' + name
    print(f'  fetch {name}')
    dest.write_bytes(http_get(url))


# --- LEWK4 legacy ---

def list_cdc_lewk4_xlsx() -> list[str]:
  html = http_get(CDC_LEWK4_INDEX).decode('utf-8', errors='replace')
  return sorted(set(re.findall(r'(lewk4_[a-z0-9_]+\.xlsx)', html, flags=re.I)))


def slug_from_filename(name: str) -> str:
  base = name.lower().replace('.xlsx', '')
  if base.startswith('lewk4_'):
    base = base[6:]
  return base


def life_table_sheet_triplet(wb) -> tuple[str, str, str]:
  """LEWK4 workbooks use total{n}/male{n}/female{n} with varying n per file."""
  pat = re.compile(r'^(male|female|total)(\d+)$', re.I)
  by_num: dict[str, dict[str, str]] = defaultdict(dict)
  for name in wb.sheetnames:
    m = pat.match(name)
    if not m:
      continue
    kind, num = m.group(1).lower(), m.group(2)
    by_num[num][kind] = name
  for num in sorted(by_num.keys(), key=lambda x: int(x)):
    b = by_num[num]
    if all(k in b for k in ('male', 'female', 'total')):
      return b['male'], b['female'], b['total']
  raise ValueError(f'No male/female/total sheet triplet in {wb.sheetnames}')


def parse_lewk4_state_workbook(path: Path) -> dict[str, list[float]]:
  wb = load_workbook(path, data_only=True, read_only=True)
  try:
    sm, sf, st = life_table_sheet_triplet(wb)
    male = dense_curve(sheet_ex_series(wb[sm]))
    female = dense_curve(sheet_ex_series(wb[sf]))
    total = dense_curve(sheet_ex_series(wb[st]))
  finally:
    wb.close()
  return {'male': male, 'female': female, 'total': total}


def collect_lewk4_state_tables(cdc_dir: Path) -> dict[str, dict[str, list[float]]]:
  files = sorted(cdc_dir.glob('lewk4_*.xlsx'))
  if not files:
    raise SystemExit(f'No lewk4_*.xlsx in {cdc_dir}. Run without --skip-cdc or check download.')

  by_code: dict[str, dict[str, list[float]]] = {}
  for path in files:
    slug = slug_from_filename(path.name)
    code = SLUG_TO_CODE.get(slug)
    if not code:
      print(f'  skip unknown file {path.name}', file=sys.stderr)
      continue
    by_code[code] = parse_lewk4_state_workbook(path)
  return by_code


def download_cdc_lewk4(cache_dir: Path) -> None:
  cache_dir.mkdir(parents=True, exist_ok=True)
  names = list_cdc_lewk4_xlsx()
  print(f'CDC LEWK4: {len(names)} workbooks listed')
  for name in names:
    dest = cache_dir / name
    if dest.exists() and dest.stat().st_size > 1000:
      print(f'  cached {name}')
      continue
    url = CDC_LEWK4_INDEX.rstrip('/') + '/' + name
    print(f'  fetch {name}')
    dest.write_bytes(http_get(url))


def fetch_world_bank_e0(base_url: str) -> dict[str, float]:
  """Country display name -> value (merges all API pages)."""
  out: dict[str, float] = {}
  page = 1
  while True:
    sep = '&' if '?' in base_url else '?'
    url = f'{base_url}{sep}page={page}'
    raw = json.loads(http_get(url))
    if not isinstance(raw, list) or len(raw) < 2:
      break
    meta, rows = raw[0], raw[1]
    if not isinstance(rows, list):
      break
    for row in rows:
      if not isinstance(row, dict):
        continue
      iso = (row.get('countryiso3code') or '').strip()
      if not iso:
        continue
      val = row.get('value')
      if val is None:
        continue
      country = row.get('country') or {}
      name = country.get('value') if isinstance(country, dict) else None
      if not name:
        continue
      try:
        out[str(name)] = float(val)
      except (TypeError, ValueError):
        continue
    pages = meta.get('pages', 1) if isinstance(meta, dict) else 1
    if page >= pages:
      break
    page += 1
  return out


def build_bundle(cdc_dir: Path, us_source: str) -> dict:
  if us_source == US_SOURCE_NVSR:
    by_code = collect_nvsr_state_tables(cdc_dir)
    period = '2022'
    source_url = 'https://www.cdc.gov/nchs/products/nvsr.htm'
    pdf_ref = 'https://www.cdc.gov/nchs/data/nvsr/nvsr74/nvsr74-12.pdf'
    data_note = (
      'State tables are NCHS U.S. State Life Tables, 2022 (NVSR vol. 74, no. 12), period life tables. '
      'Spreadsheets: CDC FTP NVSR/74-12 ({ST}1=total, {ST}2=male, {ST}3=female). '
      'International e₀ from World Bank; non-US remaining-life curves scale the US national average eₓ shape.'
    )
    ftp_note = CDC_NVSR_74_12_INDEX
  elif us_source == US_SOURCE_LEWK4:
    by_code = collect_lewk4_state_tables(cdc_dir)
    period = '1999-2001'
    source_url = 'https://www.cdc.gov/nchs/nvss/mortality/lewk4.htm'
    pdf_ref = source_url
    data_note = (
      'State tables are NCHS LEWK4 decennial period life tables (1999–2001). '
      'International e₀ from World Bank; US remaining-life uses these tables directly.'
    )
    ftp_note = CDC_LEWK4_INDEX
  else:
    raise ValueError(f'unknown us_source: {us_source}')

  if len(by_code) < 50:
    print(f'  warning: only {len(by_code)} state tables parsed', file=sys.stderr)

  male_avg = average_curves([v['male'] for v in by_code.values()])
  female_avg = average_curves([v['female'] for v in by_code.values()])
  total_avg = average_curves([v['total'] for v in by_code.values()])

  male_wb = fetch_world_bank_e0(WB_MALE)
  female_wb = fetch_world_bank_e0(WB_FEMALE)

  countries: dict[str, dict[str, float]] = {}
  for name, m in male_wb.items():
    f = female_wb.get(name)
    if f is None:
      continue
    if name == 'United States':
      continue
    countries[name] = {'maleE0': round(m, 4), 'femaleE0': round(f, 4)}

  labels = {code: US_STATE_NAMES[code] for code in by_code if code in US_STATE_NAMES}

  return {
    'schemaVersion': 2,
    'generatedAt': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
    'usTableSource': us_source,
    'cdcTablePeriod': period,
    'cdcSourceUrl': source_url,
    'cdcReportPdfUrl': pdf_ref,
    'cdcFtpSpreadsheetBaseUrl': ftp_note,
    'cdcDataNote': data_note,
    'worldBankIndicatorYear': WB_YEAR,
    'worldBankSourceUrl': 'https://data.worldbank.org/',
    'usNationalAverage': {
      'male': male_avg,
      'female': female_avg,
      'total': total_avg,
    },
    'usRegionLabels': labels,
    'usRegions': by_code,
    'countries': countries,
  }


def default_cache_dir(us_source: str) -> Path:
  if us_source == US_SOURCE_NVSR:
    return REPO_ROOT / 'scripts' / 'cache' / 'cdc_nvsr_74_12'
  return REPO_ROOT / 'scripts' / 'cache' / 'cdc_lewk4'


def main() -> None:
  parser = argparse.ArgumentParser(description='Build life-expectancy-data.json')
  parser.add_argument('--output', type=Path, default=DEFAULT_OUT)
  parser.add_argument(
    '--us-source',
    choices=(US_SOURCE_NVSR, US_SOURCE_LEWK4),
    default=US_SOURCE_NVSR,
    help=f'US state tables: {US_SOURCE_NVSR} (default, 2022) or {US_SOURCE_LEWK4} (1999–2001)',
  )
  parser.add_argument(
    '--cache-dir',
    type=Path,
    default=None,
    help='Override cache directory (default: scripts/cache/cdc_nvsr_74_12 or cdc_lewk4)',
  )
  parser.add_argument('--skip-cdc', action='store_true', help='Use existing cache only; no CDC download')
  args = parser.parse_args()

  cache_dir = args.cache_dir or default_cache_dir(args.us_source)

  if not args.skip_cdc:
    if args.us_source == US_SOURCE_NVSR:
      download_cdc_nvsr(cache_dir)
    else:
      download_cdc_lewk4(cache_dir)
  else:
    if not cache_dir.exists():
      print('--skip-cdc but cache dir missing', file=sys.stderr)
      sys.exit(1)

  bundle = build_bundle(cache_dir, args.us_source)
  args.output.parent.mkdir(parents=True, exist_ok=True)
  args.output.write_text(json.dumps(bundle, indent=2) + '\n', encoding='utf-8')
  print(f'Wrote {args.output} ({args.output.stat().st_size // 1024} KB)')


if __name__ == '__main__':
  main()
